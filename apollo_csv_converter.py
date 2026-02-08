#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
🎯 Apollo CSV to Professional Excel Converter
محوّل ملفات Apollo CSV إلى Excel احترافي

يحوّل ملفات CSV المُصدّرة من Apollo.io إلى ملف Excel منسق بالعربية
مع تنقية البيانات وإزالة Null values

Author: Data Mining Expert
Date: 2026-02-01
Version: 1.0
"""

import os
import sys
import logging
import pandas as pd
from pathlib import Path
from openpyxl.styles import PatternFill, Font, Alignment
import phonenumbers
from phonenumbers import NumberParseException

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('apollo_converter.log', encoding='utf-8'),
        logging.StreamHandler(sys.stdout)
    ]
)
logger = logging.getLogger(__name__)


class ApolloCsvConverter:
    """
    Convert Apollo.io CSV exports to professional Excel with Arabic headers
    """
    
    def __init__(self, csv_file: str):
        """
        Initialize converter
        
        Args:
            csv_file: Path to Apollo CSV file
        """
        self.csv_file = csv_file
        self.df = None
        
        if not os.path.exists(csv_file):
            raise FileNotFoundError(f"CSV file not found: {csv_file}")
        
        logger.info(f"✅ Initialized converter for: {csv_file}")
    
    def load_csv(self) -> pd.DataFrame:
        """Load and parse CSV file"""
        logger.info("\n📂 Loading CSV file...")
        
        try:
            # Try different encodings
            for encoding in ['utf-8', 'utf-8-sig', 'latin1', 'cp1252']:
                try:
                    self.df = pd.read_csv(self.csv_file, encoding=encoding)
                    logger.info(f"   ✅ Loaded with {encoding} encoding")
                    break
                except UnicodeDecodeError:
                    continue
            
            if self.df is None:
                raise ValueError("Could not read CSV with any encoding")
            
            logger.info(f"   Total rows: {len(self.df)}")
            logger.info(f"   Total columns: {len(self.df.columns)}")
            
            # Show column names
            logger.info(f"\n📋 Detected columns:")
            for i, col in enumerate(self.df.columns, 1):
                logger.info(f"   {i}. {col}")
            
            return self.df
            
        except Exception as e:
            logger.error(f"❌ Error loading CSV: {e}")
            raise
    
    def map_apollo_columns(self) -> dict:
        """
        Map Apollo column names to standard names
        
        Returns:
            Dictionary mapping Apollo columns to standard names
        """
        # Common Apollo.io column names
        column_mapping = {
            # Name fields
            'Name': 'full_name',
            'First Name': 'first_name',
            'Last Name': 'last_name',
            'Full Name': 'full_name',
            
            # Contact fields
            'Email': 'email',
            'Email Address': 'email',
            'Contact Email': 'email',
            'Phone': 'phone_number',
            'Mobile Phone': 'phone_number',
            'Direct Phone': 'phone_number',
            'Phone Number': 'phone_number',
            
            # Professional fields
            'Title': 'job_title',
            'Job Title': 'job_title',
            'Company': 'company',
            'Organization': 'company',
            'Company Name': 'company',
            
            # Location fields
            'City': 'city',
            'State': 'state',
            'Country': 'country',
            'Location': 'location',
            
            # LinkedIn
            'LinkedIn URL': 'linkedin_url',
            'LinkedIn': 'linkedin_url',
            
            # Other
            'Industry': 'industry',
            'Seniority': 'seniority',
            'Department': 'department',
            'Employees': 'company_size',
        }
        
        # Find matching columns
        mapped = {}
        for col in self.df.columns:
            for apollo_name, std_name in column_mapping.items():
                if apollo_name.lower() in col.lower():
                    mapped[col] = std_name
                    break
        
        logger.info(f"\n🔗 Column mapping:")
        for apollo_col, std_col in mapped.items():
            logger.info(f"   {apollo_col} → {std_col}")
        
        return mapped
    
    def clean_data(self) -> pd.DataFrame:
        """Clean and validate data"""
        logger.info(f"\n🧹 Data Cleaning")
        logger.info("=" * 60)
        
        initial_count = len(self.df)
        logger.info(f"   Initial records: {initial_count}")
        
        # Map columns
        column_map = self.map_apollo_columns()
        
        # Rename columns
        self.df = self.df.rename(columns=column_map)
        
        # Combine name fields if needed
        if 'full_name' not in self.df.columns:
            if 'first_name' in self.df.columns and 'last_name' in self.df.columns:
                self.df['full_name'] = self.df['first_name'].fillna('') + ' ' + self.df['last_name'].fillna('')
                self.df['full_name'] = self.df['full_name'].str.strip()
        
        # Remove duplicates
        if 'email' in self.df.columns:
            before = len(self.df)
            self.df = self.df.drop_duplicates(subset=['email'], keep='first')
            removed = before - len(self.df)
            if removed > 0:
                logger.info(f"   ✓ Removed {removed} duplicate emails")
        
        # Clean emails
        if 'email' in self.df.columns:
            # Remove invalid emails
            self.df = self.df[self.df['email'].notna()]
            self.df = self.df[self.df['email'].str.contains('@', na=False)]
            self.df['email'] = self.df['email'].str.strip().str.lower()
        
        # Format phone numbers
        if 'phone_number' in self.df.columns:
            self.df['phone_number'] = self.df['phone_number'].apply(self.format_phone)
        
        # Remove null names
        if 'full_name' in self.df.columns:
            self.df = self.df[self.df['full_name'].notna()]
            self.df = self.df[self.df['full_name'].str.strip() != '']
        
        # Combine location if needed
        if 'location' not in self.df.columns:
            location_parts = []
            for col in ['city', 'state', 'country']:
                if col in self.df.columns:
                    location_parts.append(self.df[col].fillna(''))
            
            if location_parts:
                self.df['location'] = location_parts[0]
                for part in location_parts[1:]:
                    self.df['location'] = self.df['location'] + ', ' + part
                self.df['location'] = self.df['location'].str.replace(r',\s*,', ',', regex=True)
                self.df['location'] = self.df['location'].str.strip(', ')
        
        final_count = len(self.df)
        logger.info(f"   ✅ Final clean records: {final_count}")
        logger.info(f"   ✅ Removed: {initial_count - final_count} ({(initial_count-final_count)/initial_count*100:.1f}%)")
        
        return self.df
    
    def format_phone(self, phone):
        """Format phone number to international format"""
        if pd.isna(phone) or phone == '':
            return None
        
        try:
            # Try to parse and format
            phone_str = str(phone).strip()
            
            # Try common regions
            for region in ['US', 'SA', 'AE', 'QA', 'DE', 'FR', 'GB']:
                try:
                    parsed = phonenumbers.parse(phone_str, region)
                    if phonenumbers.is_valid_number(parsed):
                        return phonenumbers.format_number(
                            parsed,
                            phonenumbers.PhoneNumberFormat.INTERNATIONAL
                        )
                except:
                    continue
            
            # If can't format, return original
            return phone_str
            
        except:
            return str(phone) if phone else None
    
    def export_to_excel(self, output_file: str = None):
        """Export to professional Excel file"""
        if output_file is None:
            base_name = Path(self.csv_file).stem
            output_file = f"{base_name}_CLEANED.xlsx"
        
        logger.info(f"\n📤 Exporting to Excel")
        logger.info("=" * 60)
        
        # Select columns to export
        export_columns = [
            'full_name', 'email', 'phone_number', 'job_title',
            'company', 'location', 'city', 'country',
            'linkedin_url', 'industry', 'seniority', 'department'
        ]
        
        # Keep only existing columns
        available_cols = [col for col in export_columns if col in self.df.columns]
        df_export = self.df[available_cols].copy()
        
        # Arabic column names
        arabic_names = {
            'full_name': 'الاسم الكامل',
            'email': 'البريد الإلكتروني',
            'phone_number': 'رقم الهاتف',
            'job_title': 'المسمى الوظيفي',
            'company': 'الشركة',
            'location': 'الموقع',
            'city': 'المدينة',
            'country': 'الدولة',
            'linkedin_url': 'رابط LinkedIn',
            'industry': 'المجال',
            'seniority': 'المستوى الوظيفي',
            'department': 'القسم'
        }
        
        df_export = df_export.rename(columns=arabic_names)
        
        # Export with formatting
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            df_export.to_excel(writer, sheet_name='Apollo Data', index=False)
            
            ws = writer.sheets['Apollo Data']
            
            # Format header
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, size=12, color="FFFFFF")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    try:
                        if cell.value and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = min(max_length + 3, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        logger.info(f"   ✅ Excel file created: {output_file}")
        
        # Create sample if large
        if len(df_export) > 200:
            sample_file = output_file.replace('.xlsx', '_SAMPLE.xlsx')
            with pd.ExcelWriter(sample_file, engine='openpyxl') as writer:
                df_export.head(200).to_excel(writer, sheet_name='Sample', index=False)
            logger.info(f"   ✅ Sample file created: {sample_file}")
        
        return output_file
    
    def generate_report(self):
        """Generate quality report"""
        print("\n" + "=" * 70)
        print("📊 CONVERSION REPORT")
        print("=" * 70)
        
        print(f"\n📈 Statistics:")
        print(f"   Total Records: {len(self.df)}")
        
        if 'email' in self.df.columns:
            email_count = self.df['email'].notna().sum()
            print(f"   With Email: {email_count} ({email_count/len(self.df)*100:.1f}%)")
        
        if 'phone_number' in self.df.columns:
            phone_count = self.df['phone_number'].notna().sum()
            print(f"   With Phone: {phone_count} ({phone_count/len(self.df)*100:.1f}%)")
        
        if 'country' in self.df.columns:
            print(f"\n🌍 Top Countries:")
            for country, count in self.df['country'].value_counts().head(5).items():
                print(f"   • {country}: {count}")
        
        if 'company' in self.df.columns:
            print(f"\n🏢 Sample Companies:")
            companies = self.df['company'].dropna().unique()[:5]
            for comp in companies:
                print(f"   • {comp}")
        
        print("\n" + "=" * 70)
        print("✅ CONVERSION COMPLETED SUCCESSFULLY!")
        print("=" * 70 + "\n")
    
    def run(self, output_file: str = None):
        """Main execution"""
        print("\n" + "=" * 70)
        print("🎯 APOLLO CSV TO EXCEL CONVERTER")
        print("   محوّل ملفات Apollo إلى Excel احترافي")
        print("=" * 70)
        
        try:
            # Load CSV
            self.load_csv()
            
            # Clean data
            self.clean_data()
            
            # Export to Excel
            output_file = self.export_to_excel(output_file)
            
            # Generate report
            self.generate_report()
            
            print(f"\n✨ Output file: {output_file}")
            print(f"✨ Ready to use for your call center!")
            
            return output_file
            
        except Exception as e:
            logger.error(f"\n❌ Error: {e}", exc_info=True)
            raise


def main():
    """Main entry point"""
    import glob
    
    print("\n🔍 Looking for Apollo CSV files in current directory...")
    
    # Find CSV files
    csv_files = glob.glob("*.csv")
    
    if not csv_files:
        print("\n❌ No CSV files found in current directory!")
        print("\n💡 Instructions:")
        print("   1. Export data from Apollo.io as CSV")
        print("   2. Place the CSV file in this directory (D:\\LLM1)")
        print("   3. Run this script again")
        print("\n   Or specify file manually:")
        print('   python apollo_csv_converter.py "your_file.csv"')
        return
    
    print(f"\n📁 Found {len(csv_files)} CSV file(s):")
    for i, f in enumerate(csv_files, 1):
        size = os.path.getsize(f) / 1024
        print(f"   {i}. {f} ({size:.1f} KB)")
    
    # If argument provided, use it
    if len(sys.argv) > 1:
        csv_file = sys.argv[1]
    # If only one file, use it
    elif len(csv_files) == 1:
        csv_file = csv_files[0]
        print(f"\n✅ Using: {csv_file}")
    # Multiple files, ask user
    else:
        print("\n❓ Which file to convert?")
        choice = input("   Enter number (or press Enter for #1): ").strip()
        
        if choice == '':
            csv_file = csv_files[0]
        else:
            try:
                idx = int(choice) - 1
                csv_file = csv_files[idx]
            except:
                print("❌ Invalid choice")
                return
    
    # Run conversion
    converter = ApolloCsvConverter(csv_file)
    converter.run()


if __name__ == "__main__":
    main()
